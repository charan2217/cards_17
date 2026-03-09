import pandas as pd
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

try:
    import streamlit as st
except Exception:
    st = None

try:
    import gspread
    from google.oauth2.service_account import Credentials
except Exception:
    gspread = None
    Credentials = None

FILE_PATH = "visiting_database.xlsx"

REQUIRED_COLUMNS = [
    "Card Holder",
    "Company Name", 
    "Designation",
    "Mobile Number",
    "Email",
    "Website",
    "Area",
    "City",
    "State",
    "Pincode",
    "Address",
    "Services/Products",
    "Other Info",
    "Date Added",
    "Status"
]

GSHEETS_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]


def _get_secret(path: str, default=None):
    if st is None:
        return default
    try:
        cur = st.secrets
        for part in path.split("."):
            cur = cur[part]
        return cur
    except Exception:
        return default


def _get_gsheets_config():
    spreadsheet_id = _get_secret("gspread.spreadsheet_id") or os.getenv("GSPREAD_SPREADSHEET_ID")
    worksheet_name = _get_secret("gspread.worksheet") or os.getenv("GSPREAD_WORKSHEET") or "Database"
    service_account_info = _get_secret("gcp_service_account")
    if not spreadsheet_id or not service_account_info:
        return None
    return {
        "spreadsheet_id": spreadsheet_id,
        "worksheet": worksheet_name,
        "service_account_info": service_account_info,
    }


def _get_worksheet():
    cfg = _get_gsheets_config()
    if cfg is None:
        return None
    if gspread is None or Credentials is None:
        raise RuntimeError("Google Sheets dependencies missing. Install gspread and google-auth")

    creds = Credentials.from_service_account_info(cfg["service_account_info"], scopes=GSHEETS_SCOPES)
    client = gspread.authorize(creds)
    sh = client.open_by_key(cfg["spreadsheet_id"])
    try:
        ws = sh.worksheet(cfg["worksheet"])
    except Exception:
        ws = sh.add_worksheet(title=cfg["worksheet"], rows=1000, cols=max(10, len(REQUIRED_COLUMNS)))
    return ws


def _read_sheet_df(ws) -> pd.DataFrame:
    values = ws.get_all_values()
    if not values:
        return pd.DataFrame(columns=REQUIRED_COLUMNS)
    header = values[0]
    rows = values[1:]
    df = pd.DataFrame(rows, columns=header)
    for col in REQUIRED_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    df = df[REQUIRED_COLUMNS]
    return df


def _ensure_sheet_header(ws):
    values = ws.get_all_values()
    if not values:
        ws.append_row(REQUIRED_COLUMNS, value_input_option="RAW")
        return

    header = values[0]
    if header != REQUIRED_COLUMNS:
        ws.clear()
        ws.append_row(REQUIRED_COLUMNS, value_input_option="RAW")
        if values[1:]:
            for row in values[1:]:
                padded = list(row) + [""] * (len(REQUIRED_COLUMNS) - len(row))
                ws.append_row(padded[: len(REQUIRED_COLUMNS)], value_input_option="RAW")


def initialize_database():
    """Initialize database with proper structure and formatting"""
    ws = _get_worksheet()
    if ws is not None:
        _ensure_sheet_header(ws)
        return "Google Sheets database ready."

    if not os.path.exists(FILE_PATH):
        # Create new database with structure
        df = pd.DataFrame(columns=REQUIRED_COLUMNS)
        df.to_excel(FILE_PATH, index=False, engine='openpyxl')
        
        # Apply formatting
        format_excel_file()
        return "Database created successfully."
    
    # Check if database needs structure update
    try:
        existing_df = pd.read_excel(FILE_PATH)
        missing_cols = [col for col in REQUIRED_COLUMNS if col not in existing_df.columns]
        
        if missing_cols:
            # Add missing columns
            for col in missing_cols:
                existing_df[col] = ""
            existing_df.to_excel(FILE_PATH, index=False, engine='openpyxl')
            format_excel_file()
            return f"Database structure updated. Added columns: {', '.join(missing_cols)}"
        
        return "Database structure verified."
    except Exception as e:
        return f"Database check failed: {str(e)}"


def format_excel_file():
    """Apply professional formatting to the Excel file"""
    try:
        wb = openpyxl.load_workbook(FILE_PATH)
        ws = wb.active
        
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Format headers
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add data validation for Status column
        if "Status" in [cell.value for cell in ws[1]]:
            status_col = None
            for i, cell in enumerate(ws[1]):
                if cell.value == "Status":
                    status_col = i + 1
                    break
            
            if status_col:
                from openpyxl.worksheet.datavalidation import DataValidation
                dv = DataValidation(type="list", formula1='"Active,Inactive,Pending"', allow_blank=True)
                ws.add_data_validation(dv)
                dv.add(f"A{status_col}:A{status_col}")
        
        wb.save(FILE_PATH)
        return True
    except Exception as e:
        print(f"Formatting error: {e}")
        return False


def clean_data(data):
    """Clean and validate data before saving"""
    cleaned_data = data.copy()
    
    # Ensure all required fields exist
    for field in REQUIRED_COLUMNS:
        if field not in cleaned_data:
            cleaned_data[field] = ""
    
    # Clean phone number
    if cleaned_data.get("Mobile Number"):
        import re
        phone = re.sub(r'[^\d+]', '', str(cleaned_data["Mobile Number"]))
        cleaned_data["Mobile Number"] = phone if len(phone) >= 10 else ""
    
    # Clean email
    if cleaned_data.get("Email"):
        import re
        email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        if not re.match(email_pattern, str(cleaned_data["Email"])):
            cleaned_data["Email"] = ""
    
    # Add metadata
    cleaned_data["Date Added"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cleaned_data["Status"] = "Active"
    
    return cleaned_data


def is_duplicate_entry(new_data, existing_df):
    """Check for duplicate entries based on multiple criteria"""
    try:
        # Check by mobile number
        if new_data.get("Mobile Number") and str(new_data["Mobile Number"]).strip():
            mobile_matches = existing_df["Mobile Number"].astype(str).str.contains(str(new_data["Mobile Number"]).strip(), na=False)
            if mobile_matches.any():
                return True, "Duplicate mobile number found"
        
        # Check by email
        if new_data.get("Email") and str(new_data["Email"]).strip():
            email_matches = existing_df["Email"].astype(str).str.contains(str(new_data["Email"]).strip(), na=False)
            if email_matches.any():
                return True, "Duplicate email found"
        
        # Check by card holder + company combination
        if (new_data.get("Card Holder") and new_data.get("Company Name") and 
            str(new_data["Card Holder"]).strip() and str(new_data["Company Name"]).strip()):
            holder_matches = existing_df["Card Holder"].astype(str).str.contains(str(new_data["Card Holder"]).strip(), na=False)
            company_matches = existing_df["Company Name"].astype(str).str.contains(str(new_data["Company Name"]).strip(), na=False)
            if (holder_matches & company_matches).any():
                return True, "Duplicate card holder + company combination found"
        
        return False, ""
    except Exception as e:
        print(f"Duplicate check error: {e}")
        return False, ""


def save_to_database(data):
    """Enhanced database save function with validation and formatting"""
    try:
        # Initialize database if needed
        init_status = initialize_database()
        
        # Clean and validate data
        cleaned_data = clean_data(data)
        
        ws = _get_worksheet()
        if ws is not None:
            _ensure_sheet_header(ws)
            existing_df = _read_sheet_df(ws)
        else:
            existing_df = pd.read_excel(FILE_PATH)
        
        # Check for duplicates
        is_duplicate, duplicate_reason = is_duplicate_entry(cleaned_data, existing_df)
        if is_duplicate:
            return f"⚠️ Duplicate Entry: {duplicate_reason}"
        
        # Create new row
        new_df = pd.DataFrame([cleaned_data])
        
        # Append data
        updated_df = pd.concat([existing_df, new_df], ignore_index=True)
        
        if ws is not None:
            row = [str(cleaned_data.get(col, "")) for col in REQUIRED_COLUMNS]
            ws.append_row(row, value_input_option="RAW")
        else:
            # Save to Excel
            updated_df.to_excel(FILE_PATH, index=False, engine='openpyxl')
            # Apply formatting
            format_excel_file()
        
        return f"✅ Successfully added: {cleaned_data.get('Card Holder', 'Unknown')} - {cleaned_data.get('Company Name', 'Unknown')}"
        
    except PermissionError:
        return "❌ Please close the Excel file before saving."
    except Exception as e:
        return f"❌ Error saving to database: {str(e)}"


def get_database_stats():
    """Get database statistics"""
    try:
        ws = _get_worksheet()
        if ws is not None:
            _ensure_sheet_header(ws)
            df = _read_sheet_df(ws)
        else:
            df = pd.read_excel(FILE_PATH)
        total_entries = len(df)
        active_entries = len(df[df["Status"] == "Active"]) if "Status" in df.columns else total_entries
        
        return {
            "total_entries": total_entries,
            "active_entries": active_entries,
            "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
    except Exception as e:
        return {"error": str(e)}


def export_database_backup():
    """Create a backup of the database"""
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = f"visiting_database_backup_{timestamp}.xlsx"

        ws = _get_worksheet()
        if ws is not None:
            _ensure_sheet_header(ws)
            df = _read_sheet_df(ws)
        else:
            df = pd.read_excel(FILE_PATH)

        df.to_excel(backup_path, index=False, engine='openpyxl')
        
        return f"✅ Backup created: {backup_path}"
    except Exception as e:
        return f"❌ Backup failed: {str(e)}"